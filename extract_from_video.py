import sys
import os
import time
import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from sleeping_pose import PersonPostureTracker
except ImportError:
    print("[ERROR] sleeping_pose.py not found. Place it in the same folder.")
    sys.exit(1)

try:
    from ultralytics import YOLO
except ImportError:
    print("[ERROR] ultralytics not installed.  pip install ultralytics")
    sys.exit(1)

# =============================================================================
# CONFIG  --  change these values each time you process a new batch
# =============================================================================
INPUT_FOLDER   = r"D:\Video Tido"                              # folder with all videos
PERSON_ID      = "P02"                                         # person identifier
MODEL_PATH     = "yolo11m-pose.pt"
RESULTS_ROOT   = r"D:\Sleeping\Results"
PROCESSED_LOG  = os.path.join(r"D:\Sleeping\Results", "processed_videos.txt")

IMG_SIZE       = 640
CONF_THRESHOLD = 0.45
SAVE_EVERY_N   = 40          # 1 in every N clean frames saved to dataset
JPEG_QUALITY   = 95
OUTPUT_SIZE    = None        # None = keep source resolution
LABELS         = ["Supine", "Lateral_Left", "Lateral_Right", "Prone", "No_Person"]


# =============================================================================
# EXCEL STYLES
# =============================================================================
FILL_YELLOW = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
FILL_HEADER = PatternFill("solid", start_color="2C3E50", end_color="2C3E50")
FILL_EVEN   = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
FILL_LABEL  = {
    "Supine"        : PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3"),
    "Lateral_Left"  : PatternFill("solid", start_color="D6EAF8", end_color="D6EAF8"),
    "Lateral_Right" : PatternFill("solid", start_color="D2E9F7", end_color="D2E9F7"),
    "Prone"         : PatternFill("solid", start_color="FDEBD0", end_color="FDEBD0"),
    "No_Person"     : PatternFill("solid", start_color="EAECEE", end_color="EAECEE"),
    "Unknown"       : PatternFill("solid", start_color="F2F3F4", end_color="F2F3F4"),
}
FONT_HEADER = Font(bold=True, color="FFFFFF", name="Arial", size=9)
FONT_NORMAL = Font(name="Arial", size=9)
FONT_FLAG   = Font(name="Arial", size=9, bold=True)
THIN_BORDER = Border(
    left  =Side(style="thin", color="CCCCCC"),
    right =Side(style="thin", color="CCCCCC"),
    top   =Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
LABEL_COLOURS = {
    "Supine"        : (0,   255, 180),
    "Lateral_Left"  : (0,   180, 255),
    "Lateral_Right" : (0,   220, 255),
    "Prone"         : (255, 140,   0),
    "No_Person"     : (160, 160, 160),
    "Unknown"       : (160, 160, 160),
}
KP_NAMES = [
    "nose",
    "left_eye",  "right_eye",
    "left_ear",  "right_ear",
    "left_shoulder", "right_shoulder",
    "left_elbow",    "right_elbow",
    "left_wrist",    "right_wrist",
    "left_hip",      "right_hip",
    "left_knee",     "right_knee",
    "left_ankle",    "right_ankle",
]


# =============================================================================
# PROCESSED LOG
# =============================================================================
def _load_processed_log():
    """Load set of already processed video filenames."""
    if not os.path.isfile(PROCESSED_LOG):
        return set()
    with open(PROCESSED_LOG, "r") as f:
        return set(line.strip() for line in f if line.strip())

def _save_processed_log(filename):
    """Append a completed video filename to the log."""
    os.makedirs(os.path.dirname(PROCESSED_LOG), exist_ok=True)
    with open(PROCESSED_LOG, "a") as f:
        f.write(filename + "\n")

def _get_next_session_number():
    """Get next session number based on existing result folders."""
    if not os.path.isdir(RESULTS_ROOT):
        return 1
    existing = [
        d for d in os.listdir(RESULTS_ROOT)
        if os.path.isdir(os.path.join(RESULTS_ROOT, d))
    ]
    return len(existing) + 1


# =============================================================================
# KEYPOINT HELPERS
# =============================================================================
def _body_angle(kp):
    ls, rs = kp[5], kp[6]
    lh, rh = kp[11], kp[12]
    if ls[2] < 0.5 or rs[2] < 0.5: return None
    if lh[2] < 0.5 and rh[2] < 0.5: return None
    smid = ((ls[0]+rs[0])/2, (ls[1]+rs[1])/2)
    if lh[2] > 0.5 and rh[2] > 0.5:
        hmid = ((lh[0]+rh[0])/2, (lh[1]+rh[1])/2)
    elif lh[2] > 0.5:
        hmid = (lh[0], lh[1])
    else:
        hmid = (rh[0], rh[1])
    dx = hmid[0]-smid[0]; dy = hmid[1]-smid[1]
    return round(float(np.degrees(np.arctan2(abs(dy), abs(dx)+1e-6))), 2)

def _shoulder_asym(kp):
    ls, rs = kp[5], kp[6]
    if ls[2] < 0.5 or rs[2] < 0.5: return None
    return round(abs(ls[1]-rs[1]) / max(abs(ls[0]-rs[0]), 20.0), 4)

def _ear_asym(kp):
    le, re = kp[3], kp[4]
    if le[2] < 0.28 or re[2] < 0.28: return None
    return round(abs(le[1]-re[1]) / max(abs(le[0]-re[0]), 10.0), 4)

def _movement(prev_kp, curr_kp):
    if prev_kp is None: return 0.0
    vis = (curr_kp[:,2]>0.5) & (prev_kp[:,2]>0.5)
    if vis.sum() == 0: return 0.0
    return round(float(np.linalg.norm(
        curr_kp[vis,:2]-prev_kp[vis,:2], axis=1).mean()), 3)


# =============================================================================
# DIRECTORY SETUP
# =============================================================================
def _make_session_dirs(session_root):
    dirs = {}
    for lbl in LABELS + ["Unknown"]:
        p = os.path.join(session_root, "images", "dataset", lbl)
        os.makedirs(p, exist_ok=True)
        dirs[("dataset", lbl)] = p
    p = os.path.join(session_root, "images", "review", "transitions")
    os.makedirs(p, exist_ok=True)
    dirs[("review", "transitions")] = p
    return dirs


# =============================================================================
# OVERLAY / IMAGE HELPERS
# =============================================================================
def _burn_overlay(frame, label, confidence, flag_str, prev_label=None):
    img    = frame.copy()
    colour = LABEL_COLOURS.get(label, (200, 200, 200))
    font   = cv2.FONT_HERSHEY_SIMPLEX
    txt    = f"{label.replace('_',' ')}  {confidence:.0%}"
    cv2.putText(img, txt, (10,34), font, 0.85, (0,0,0),  4, cv2.LINE_AA)
    cv2.putText(img, txt, (10,34), font, 0.85, colour,    2, cv2.LINE_AA)
    if prev_label and prev_label != label:
        arrow = f"{prev_label} -> {label}"
        cv2.putText(img, arrow, (10,62), font, 0.55, (0,0,0),     3, cv2.LINE_AA)
        cv2.putText(img, arrow, (10,62), font, 0.55, (0,255,255), 1, cv2.LINE_AA)
    if flag_str:
        (tw,_),_ = cv2.getTextSize(flag_str, font, 0.5, 1)
        x = img.shape[1]-tw-14
        cv2.rectangle(img, (x-4,8), (x+tw+4,26), (0,0,0), -1)
        cv2.putText(img, flag_str, (x,22), font, 0.5, (0,255,255), 1, cv2.LINE_AA)
    return img

def _save_image(img, path):
    out = cv2.resize(img, OUTPUT_SIZE, interpolation=cv2.INTER_AREA) if OUTPUT_SIZE else img
    cv2.imwrite(path, out, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])


# =============================================================================
# EXCEL HELPERS
# =============================================================================
def _build_header():
    base = ["person_id","frame","timestamp_s","video_time","image_file"]
    for name in KP_NAMES:
        base += [f"{name}_x", f"{name}_y", f"{name}_conf"]
    base += ["body_angle","shoulder_asym","ear_asym","movement",
             "auto_label","confidence","flag","verified_label","notes"]
    return base

def _col_widths(headers):
    w = {}
    for h in headers:
        if h == "person_id":                       w[h] = 10
        elif h == "frame":                         w[h] = 7
        elif h in ("timestamp_s","video_time"):    w[h] = 12
        elif h == "image_file":                    w[h] = 44
        elif h.endswith(("_x","_y","_conf")):      w[h] = 8
        elif h in ("body_angle","shoulder_asym",
                   "ear_asym","movement"):         w[h] = 13
        elif h in ("auto_label","verified_label"): w[h] = 16
        elif h == "confidence":                    w[h] = 11
        elif h == "flag":                          w[h] = 24
        elif h == "notes":                         w[h] = 25
        else:                                      w[h] = 10
    return w

def _write_sheet(ws, headers, rows, title="Data"):
    ws.title = title; ws.freeze_panes = "A2"
    col_w     = _col_widths(headers)
    label_col = headers.index("auto_label") + 1
    flag_col  = headers.index("flag")       + 1
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font=FONT_HEADER; cell.fill=FILL_HEADER
        cell.alignment=Alignment(horizontal="center", vertical="center")
        cell.border=THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col_w.get(h, 10)
    ws.row_dimensions[1].height = 20
    for ri, row in enumerate(rows, 2):
        fv = row[flag_col-1]; lv = row[label_col-1]
        is_t  = "TRANSITION" in str(fv)
        rfill = FILL_YELLOW if is_t else \
                (FILL_EVEN  if ri%2==0 else None)
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border=THIN_BORDER
            cell.alignment=Alignment(horizontal="center", vertical="center")
            if rfill:
                cell.fill=rfill
                cell.font=FONT_FLAG if is_t else FONT_NORMAL
            else:
                cell.font=FONT_NORMAL
                if ci==label_col and lv in FILL_LABEL:
                    cell.fill=FILL_LABEL[lv]
        ws.row_dimensions[ri].height = 15
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def _write_summary(wb, video_path, headers, all_rows, flagged_rows,
                   img_dir, person_id, session, tag):
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 44
    lc = headers.index("auto_label"); fc = headers.index("flag")
    counts = {}
    for r in all_rows:
        counts[r[lc]] = counts.get(r[lc], 0)+1
    data = [
        ("Person ID",          person_id),
        ("Session",            session),
        ("Session tag",        tag),
        ("Video file",         os.path.basename(video_path)),
        ("Total frames",       len(all_rows)),
        ("Flagged for review", len(flagged_rows)),
        ("Transitions",        sum(1 for r in all_rows if "TRANSITION" in str(r[fc]))),
        ("Images saved to",    img_dir),
        ("",""),
        ("Label distribution", ""),
    ]
    for lbl, cnt in sorted(counts.items()):
        pct = cnt/len(all_rows)*100 if all_rows else 0
        data.append((f"  {lbl}", f"{cnt} frames  ({pct:.1f}%)"))
    data += [
        ("",""),
        ("HOW TO USE",""),
        ("YELLOW rows",        "Label changed — verify posture is correct"),
        ("verified_label col", "Type correct label here if auto_label is wrong"),
        ("Valid labels",       "Supine | Lateral_Left | Lateral_Right | Prone | No_Person"),
        ("",""),
        ("NEXT STEPS",""),
        ("1. Review flagged rows", "Open _review.xlsx, fix verified_label column"),
        ("2. Save the Excel",      "Save after all corrections are done"),
        ("3. Combine manually",    "Copy verified Excel to your combined/keypoints folder"),
        ("4. Train",               "Run train_xgboost.py / train_lstm.py / train_hybrid.py"),
    ]
    for ri,(k,v) in enumerate(data,1):
        ck=ws.cell(row=ri,column=1,value=k); cv=ws.cell(row=ri,column=2,value=v)
        if k in ("HOW TO USE","Label distribution","NEXT STEPS"):
            ck.font=Font(bold=True,name="Arial",size=10,color="2C3E50")
        elif k.startswith("  ") or k.startswith("1.") or k.startswith("2.") \
             or k.startswith("3.") or k.startswith("4."):
            ck.font=Font(name="Arial",size=9,italic=True)
            cv.font=Font(name="Arial",size=9)
        else:
            ck.font=Font(bold=True,name="Arial",size=9)
            cv.font=Font(name="Arial",size=9)
    cr=ri+2
    ws.cell(row=cr,column=1,value="Colour key").font=Font(bold=True,name="Arial",size=9)
    c=ws.cell(row=cr+1,column=1,value="Transition")
    c.fill=FILL_YELLOW; c.font=Font(name="Arial",size=9,bold=True)
    ws.title="Summary"


# =============================================================================
# CORE EXTRACTION
# =============================================================================
def extract_video(video_path, model, person_id, session, tag, session_root):
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print(f"  [ERROR] Cannot open: {video_path}"); return False

    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps_video    = cap.get(cv2.CAP_PROP_FPS) or 25.0
    print(f"  Frames: {total_frames}  |  FPS: {fps_video:.1f}")

    sess_dirs    = _make_session_dirs(session_root)
    date_compact = tag[1:11].replace("-","")
    sess_num     = "".join(filter(str.isdigit, session)) or "1"
    img_prefix   = f"{person_id}_{date_compact}_s{sess_num}"

    tracker      = PersonPostureTracker(person_idx=0, smooth=True)
    headers      = _build_header()
    all_rows     = []
    prev_kp      = None
    prev_label   = None
    frame_idx    = 0
    saved_counts = {"dataset":0, "review":0}

    while True:
        ret, frame = cap.read()
        if not ret: break

        frame_idx += 1
        timestamp  = frame_idx / fps_video
        m,s = divmod(timestamp,60); h,m = divmod(m,60)
        video_time = f"{int(h):02d}:{int(m):02d}:{s:05.2f}"

        infer   = cv2.resize(frame, (IMG_SIZE, IMG_SIZE))
        results = model.predict(infer, imgsz=IMG_SIZE, conf=CONF_THRESHOLD,
                                device="cuda", verbose=False)[0]

        if (results.keypoints is not None and results.boxes is not None
                and len(results.boxes) > 0):
            kp_data = results.keypoints.data.cpu().numpy()[0]
            result  = tracker.update(kp_data, timestamp)
            mov     = _movement(prev_kp, kp_data)
            prev_kp = kp_data.copy()
        else:
            kp_data = np.zeros((17,3), dtype=np.float32)
            result  = tracker.update_no_person(timestamp)
            mov     = 0.0; prev_kp = None

        label      = result.posture
        confidence = round(result.confidence, 4)
        b_angle    = _body_angle(kp_data)
        sh_asym    = _shoulder_asym(kp_data)
        ear_asym_v = _ear_asym(kp_data)

        is_transition = prev_label is not None and label != prev_label
        flag_str      = "TRANSITION" if is_transition else ""

        clean     = frame.copy(); clean[0:45, 0:160] = 0
        annotated = results.plot(img=clean, line_width=1, kpt_radius=3,
                                 labels=False, boxes=False, kpt_line=True)
        overlay   = _burn_overlay(annotated, label, confidence, flag_str,
                                  prev_label if is_transition else None)
        conf_pct  = int(confidence*100)
        img_filename = ""

        if is_transition:
            sp    = (prev_label or "None").replace(" ","_")
            fname = f"{img_prefix}_f{frame_idx:06d}_{sp}_to_{label}.jpg"
            fpath = os.path.join(sess_dirs[("review","transitions")], fname)
            _save_image(overlay, fpath)
            img_filename = os.path.join("images","review","transitions",fname)
            saved_counts["review"] += 1

        else:
            if frame_idx % SAVE_EVERY_N == 0:
                lbl_safe = label.replace(" ","_")
                fname    = f"{img_prefix}_f{frame_idx:06d}_{lbl_safe}_{conf_pct}pct.jpg"
                fpath    = os.path.join(sess_dirs[("dataset", label)], fname)
                _save_image(overlay, fpath)
                img_filename = os.path.join("images","dataset",label,fname)
                saved_counts["dataset"] += 1

        prev_label = label

        row = [person_id, frame_idx, round(timestamp,3), video_time, img_filename]
        for i in range(17):
            row += [round(float(kp_data[i,0]),2),
                    round(float(kp_data[i,1]),2),
                    round(float(kp_data[i,2]),4)]
        row += [b_angle, sh_asym, ear_asym_v, mov,
                label, confidence, flag_str, "", ""]
        all_rows.append(row)

        if frame_idx % 500 == 0:
            print(f"  {frame_idx}/{total_frames} | "
                  f"dataset: {saved_counts['dataset']} | "
                  f"review: {saved_counts['review']}")

    cap.release()
    flagged_rows = [r for r in all_rows if r[headers.index("flag")] != ""]
    print(f"  Done — {len(all_rows)} frames | {len(flagged_rows)} flagged | "
          f"{saved_counts['dataset']} dataset | {saved_counts['review']} review")

    # ── Excel ──────────────────────────────────────────────────────────────────
    out_full = os.path.join(session_root, f"{tag}_labels.xlsx")
    out_rev  = os.path.join(session_root, f"{tag}_review.xlsx")

    wb = Workbook()
    _write_summary(wb, video_path, headers, all_rows, flagged_rows,
                   os.path.join(session_root,"images"), person_id, session, tag)
    _write_sheet(wb.create_sheet("All Frames"),   headers, all_rows,     title="All Frames")
    _write_sheet(wb.create_sheet("Flagged Only"), headers, flagged_rows, title="Flagged Only")
    wb.active = wb["Summary"]
    wb.save(out_full)
    print(f"  Labels:  {out_full}")

    wb2 = Workbook()
    _write_summary(wb2, video_path, headers, all_rows, flagged_rows,
                   os.path.join(session_root,"images"), person_id, session, tag)
    ws_rev = wb2.create_sheet("Review")
    _write_sheet(ws_rev, headers, flagged_rows, title="Review")
    wb2.active = wb2["Review"]
    if "Sheet" in wb2.sheetnames: del wb2["Sheet"]
    wb2.save(out_rev)
    print(f"  Review:  {out_rev}")

    return True


# =============================================================================
# MAIN
# =============================================================================
def main():
    print(f"Person  : {PERSON_ID}")
    print(f"Folder  : {INPUT_FOLDER}")
    print(f"Results : {RESULTS_ROOT}")

    if not os.path.isdir(INPUT_FOLDER):
        print(f"[ERROR] Folder not found: {INPUT_FOLDER}")
        sys.exit(1)

    # Get all videos sorted by name
    all_videos = sorted([
        os.path.join(INPUT_FOLDER, f)
        for f in os.listdir(INPUT_FOLDER)
        if f.lower().endswith((".mp4", ".avi", ".mov", ".mkv"))
    ])

    if not all_videos:
        print("[ERROR] No video files found in folder.")
        sys.exit(1)

    # Filter out already processed videos
    processed = _load_processed_log()
    pending   = [v for v in all_videos
                 if os.path.basename(v) not in processed]

    print(f"\nTotal videos  : {len(all_videos)}")
    print(f"Already done  : {len(processed)}")
    print(f"To process    : {len(pending)}")

    if not pending:
        print("\nAll videos already processed. Nothing to do.")
        sys.exit(0)

    # Check GPU
    try:
        import torch
        if not torch.cuda.is_available():
            print("[ERROR] CUDA not available. Check your PyTorch installation.")
            sys.exit(1)
        gpu_name = torch.cuda.get_device_name(0)
        vram_gb  = torch.cuda.get_device_properties(0).total_memory / 1e9
        print(f"\nGPU : {gpu_name}  ({vram_gb:.1f} GB VRAM)")
    except ImportError:
        print("[ERROR] torch not installed. pip install torch")
        sys.exit(1)

    # Load model once, reuse for all videos
    model = YOLO(MODEL_PATH)
    model.to("cuda")
    print(f"Model : {MODEL_PATH} loaded on GPU\n")

    # Process each pending video
    for video_path in pending:
        session_num  = _get_next_session_number()
        session_tag  = f"session{session_num}"
        video_name   = os.path.splitext(os.path.basename(video_path))[0]
        date_part    = video_name[:10] if len(video_name) >= 10 else video_name
        tag          = f"({PERSON_ID}_{date_part}_{session_tag})"
        folder_name  = f"{PERSON_ID}_{date_part}_{session_tag}"
        session_root = os.path.join(RESULTS_ROOT, folder_name)
        os.makedirs(session_root, exist_ok=True)

        print(f"[{session_num}] {os.path.basename(video_path)}")
        print(f"     Output → {session_root}")
        t0 = time.time()

        ok = extract_video(video_path, model,
                           person_id=PERSON_ID,
                           session=session_tag,
                           tag=tag,
                           session_root=session_root)

        if ok:
            # Log as done — will be skipped on next run
            _save_processed_log(os.path.basename(video_path))
            print(f"     Done in {time.time()-t0:.1f}s  ✓ logged\n")
        else:
            print(f"     [FAILED] Not logged — will retry on next run\n")

    print(f"All done. Results → {RESULTS_ROOT}")
    print(f"Log file → {PROCESSED_LOG}")


if __name__ == "__main__":
    main()