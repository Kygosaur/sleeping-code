"""
train_lstm.py  --  Train LSTM classifier on sleep posture keypoint sequences.

Standalone script. Change the CONFIG section at the top, then run:
    python train_lstm.py

Outputs (saved to OUTPUT_DIR):
    lstm_model.pt              -- trained model, reusable
    lstm_results.xlsx          -- per-class F1, confusion matrix, loss curve chart
"""

# =============================================================================
# CONFIG  --  change these paths each time you use a new dataset
# =============================================================================
DATASET_DIR   = r"C:\SleepPosture\dataset"       # folder containing all (*_labels.xlsx) files
OUTPUT_DIR    = r"C:\SleepPosture\results\LSTM Training"      # where model + results are saved
SHEET_NAME    = "All Frames"                      # sheet inside each Excel to read
TEST_PERSONS  = ["P05"]                           # person IDs held out for testing
TRAIN_PERSONS = ["P01", "P02", "P03", "P04"]     # person IDs used for training
                                                  # set both to [] to auto split 80/20

# LSTM hyperparameters
SEQUENCE_LEN  = 30       # frames per sequence (30 frames = 3 sec at 10fps)
HIDDEN_SIZE   = 128      # LSTM hidden units
NUM_LAYERS    = 2        # stacked LSTM layers
DROPOUT       = 0.3      # dropout between layers
EPOCHS        = 50       # training epochs
BATCH_SIZE    = 64       # samples per batch
LEARNING_RATE = 0.001    # Adam learning rate
# =============================================================================

import os
import sys
import glob
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

try:
    import torch
    import torch.nn as nn
    from torch.utils.data import DataLoader, TensorDataset
except ImportError:
    print("[ERROR] PyTorch not installed. Run: pip install torch")
    sys.exit(1)

from sklearn.metrics import (classification_report, confusion_matrix,
                             accuracy_score, f1_score)
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split


# =============================================================================
# CONSTANTS
# =============================================================================
LABELS = ["Supine", "Lateral_Left", "Lateral_Right", "Prone", "No_Person"]

FEATURE_SUFFIXES = ["_x", "_y", "_conf"]
DERIVED_FEATURES = ["body_angle", "shoulder_asym", "ear_asym", "movement"]

FILL_HEADER = PatternFill("solid", start_color="2C3E50", end_color="2C3E50")
FILL_EVEN   = PatternFill("solid", start_color="F2F3F4", end_color="F2F3F4")
FILL_GOOD   = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")
FILL_WARN   = PatternFill("solid", start_color="FDEBD0", end_color="FDEBD0")
FILL_BAD    = PatternFill("solid", start_color="FADBD8", end_color="FADBD8")
FILL_LABEL  = {
    "Supine"        : PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3"),
    "Lateral_Left"  : PatternFill("solid", start_color="D6EAF8", end_color="D6EAF8"),
    "Lateral_Right" : PatternFill("solid", start_color="D2E9F7", end_color="D2E9F7"),
    "Prone"         : PatternFill("solid", start_color="FDEBD0", end_color="FDEBD0"),
    "No_Person"     : PatternFill("solid", start_color="EAECEE", end_color="EAECEE"),
}
FONT_HEADER = Font(bold=True, color="FFFFFF", name="Arial", size=9)
FONT_BOLD   = Font(bold=True, name="Arial", size=9)
FONT_NORMAL = Font(name="Arial", size=9)
FONT_TITLE  = Font(bold=True, name="Arial", size=12, color="2C3E50")
THIN        = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)
CENTER      = Alignment(horizontal="center", vertical="center")


# =============================================================================
# LSTM MODEL
# =============================================================================
class SleepLSTM(nn.Module):
    def __init__(self, input_size, hidden_size, num_layers, num_classes, dropout):
        super().__init__()
        self.lstm = nn.LSTM(
            input_size, hidden_size, num_layers,
            batch_first=True, dropout=dropout if num_layers > 1 else 0.0
        )
        self.dropout = nn.Dropout(dropout)
        self.fc      = nn.Linear(hidden_size, num_classes)

    def forward(self, x):
        out, _ = self.lstm(x)
        out     = self.dropout(out[:, -1, :])   # last time step
        return self.fc(out)


# =============================================================================
# DATA LOADING
# =============================================================================
def load_dataset(dataset_dir, sheet_name):
    files = glob.glob(os.path.join(dataset_dir, "*_labels.xlsx"))
    if not files:
        print(f"[ERROR] No *_labels.xlsx files found in: {dataset_dir}")
        sys.exit(1)

    print(f"\nFound {len(files)} Excel file(s):")
    dfs = []
    for f in files:
        print(f"  Loading: {os.path.basename(f)}")
        df = pd.read_excel(f, sheet_name=sheet_name)
        dfs.append(df)

    master = pd.concat(dfs, ignore_index=True)
    print(f"  Total rows loaded: {len(master)}")
    return master


def prepare_data(master):
    if "verified_label" in master.columns:
        master["label"] = master["verified_label"].astype(str).str.strip()
        master["label"] = master["label"].replace({"": None, "nan": None, "None": None})
        master["label"] = master["label"].combine_first(master["auto_label"])
    else:
        master["label"] = master["auto_label"]

    master = master[master["label"].isin(LABELS)].copy()

    if "flag" in master.columns and "verified_label" in master.columns:
        unverified = (master["flag"].astype(str).str.strip() != "") & \
                     (master["verified_label"].astype(str).str.strip().isin(["", "nan", "None"]))
        dropped = unverified.sum()
        master  = master[~unverified].copy()
        if dropped:
            print(f"  Dropped {dropped} unverified flagged rows")

    feat_cols = [c for c in master.columns
                 if any(c.endswith(s) for s in FEATURE_SUFFIXES)
                 or c in DERIVED_FEATURES]
    feat_cols = [c for c in feat_cols if c in master.columns]
    master[feat_cols] = master[feat_cols].fillna(0)

    print(f"  Clean samples: {len(master)} | Features: {len(feat_cols)}")
    print(f"  Label distribution:\n{master['label'].value_counts().to_string()}")
    return master, feat_cols


def make_sequences(df, feat_cols, le, seq_len):
    """Slide a window of seq_len frames, label = last frame's label."""
    X_list, y_list = [], []
    arr    = df[feat_cols].values.astype(np.float32)
    labels = le.transform(df["label"].values)

    for i in range(len(arr) - seq_len + 1):
        X_list.append(arr[i : i + seq_len])
        y_list.append(labels[i + seq_len - 1])

    return np.array(X_list, dtype=np.float32), np.array(y_list, dtype=np.int64)


def split_and_sequence(master, feat_cols):
    le = LabelEncoder()
    le.fit(LABELS)

    if TRAIN_PERSONS and TEST_PERSONS and "person_id" in master.columns:
        train_df = master[master["person_id"].isin(TRAIN_PERSONS)]
        test_df  = master[master["person_id"].isin(TEST_PERSONS)]
        print(f"\n  Train persons: {TRAIN_PERSONS} ({len(train_df)} rows)")
        print(f"  Test persons:  {TEST_PERSONS} ({len(test_df)} rows)")
    else:
        train_df, test_df = train_test_split(master, test_size=0.2,
                                             random_state=42, stratify=master["label"])

    X_train, y_train = make_sequences(train_df, feat_cols, le, SEQUENCE_LEN)
    X_test,  y_test  = make_sequences(test_df,  feat_cols, le, SEQUENCE_LEN)

    print(f"  Train sequences: {len(X_train)} | Test sequences: {len(X_test)}")
    return X_train, y_train, X_test, y_test, le


# =============================================================================
# TRAINING
# =============================================================================
def train_lstm(X_train, y_train, input_size, num_classes):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"\nTraining LSTM on {device} ...")

    model = SleepLSTM(input_size, HIDDEN_SIZE, NUM_LAYERS, num_classes, DROPOUT).to(device)

    # class weights for imbalance
    class_counts = np.bincount(y_train, minlength=num_classes).astype(np.float32)
    class_counts = np.where(class_counts == 0, 1, class_counts)
    weights      = torch.tensor(1.0 / class_counts).to(device)

    criterion = nn.CrossEntropyLoss(weight=weights)
    optimizer = torch.optim.Adam(model.parameters(), lr=LEARNING_RATE)
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=15, gamma=0.5)

    X_t = torch.tensor(X_train)
    y_t = torch.tensor(y_train)
    loader = DataLoader(TensorDataset(X_t, y_t), batch_size=BATCH_SIZE, shuffle=True)

    history = []
    for epoch in range(1, EPOCHS + 1):
        model.train()
        epoch_loss = 0.0
        correct    = 0
        total      = 0
        for xb, yb in loader:
            xb, yb = xb.to(device), yb.to(device)
            optimizer.zero_grad()
            out  = model(xb)
            loss = criterion(out, yb)
            loss.backward()
            optimizer.step()
            epoch_loss += loss.item() * len(xb)
            correct    += (out.argmax(1) == yb).sum().item()
            total      += len(xb)
        scheduler.step()
        avg_loss = epoch_loss / total
        acc      = correct / total
        history.append((epoch, round(avg_loss, 4), round(acc, 4)))
        if epoch % 10 == 0 or epoch == 1:
            print(f"  Epoch {epoch:3d}/{EPOCHS}  loss={avg_loss:.4f}  acc={acc:.4f}")

    return model, device, history


# =============================================================================
# EVALUATION
# =============================================================================
def evaluate(model, le, X_test, y_test, device):
    model.eval()
    with torch.no_grad():
        X_t    = torch.tensor(X_test).to(device)
        logits = model(X_t)
        y_pred = logits.argmax(1).cpu().numpy()

    y_test_labels = le.inverse_transform(y_test)
    y_pred_labels = le.inverse_transform(y_pred)

    present_labels = sorted(set(y_test_labels) | set(y_pred_labels),
                            key=lambda x: LABELS.index(x) if x in LABELS else 99)

    report   = classification_report(y_test_labels, y_pred_labels,
                                     labels=present_labels, output_dict=True)
    cm       = confusion_matrix(y_test_labels, y_pred_labels, labels=present_labels)
    acc      = accuracy_score(y_test_labels, y_pred_labels)
    macro_f1 = f1_score(y_test_labels, y_pred_labels,
                        average="macro", labels=present_labels)

    print(f"\n  Accuracy:  {acc:.4f}")
    print(f"  Macro F1:  {macro_f1:.4f}")
    print("\n" + classification_report(y_test_labels, y_pred_labels, labels=present_labels))
    return report, cm, present_labels, acc, macro_f1


# =============================================================================
# EXCEL RESULTS
# =============================================================================
def _hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_HEADER; c.fill = FILL_HEADER
    c.alignment = CENTER; c.border = THIN
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row, col, value, bold=False, fill=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = FONT_BOLD if bold else FONT_NORMAL
    c.alignment = CENTER; c.border = THIN
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    return c


def write_results_excel(output_path, report, cm, present_labels,
                        acc, macro_f1, history, feat_cols):
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    ws["A1"] = "LSTM — Sleep Posture Classification"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = CENTER

    rows = [
        ("Model",           "LSTM"),
        ("Sequence Length",  SEQUENCE_LEN),
        ("Hidden Size",      HIDDEN_SIZE),
        ("Num Layers",       NUM_LAYERS),
        ("Dropout",          DROPOUT),
        ("Epochs",           EPOCHS),
        ("Batch Size",       BATCH_SIZE),
        ("Learning Rate",    LEARNING_RATE),
        ("Train Persons",    ", ".join(TRAIN_PERSONS) if TRAIN_PERSONS else "Auto 80%"),
        ("Test Persons",     ", ".join(TEST_PERSONS)  if TEST_PERSONS  else "Auto 20%"),
        ("Features",         len(feat_cols)),
        ("Overall Accuracy", f"{acc:.4f}"),
        ("Macro F1",         f"{macro_f1:.4f}"),
    ]
    for ri, (k, v) in enumerate(rows, start=3):
        ws.cell(row=ri, column=1, value=k).font = FONT_BOLD
        ws.cell(row=ri, column=2, value=v).font = FONT_NORMAL
        ws.row_dimensions[ri].height = 16

    # ── Sheet 2: Per-Class Metrics ──────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Class Metrics")
    headers = ["Class", "Precision", "Recall", "F1-Score", "Support"]
    widths  = [18, 13, 13, 13, 13]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        _hdr(ws2, 1, ci, h, width=w)
    ws2.row_dimensions[1].height = 20

    for ri, lbl in enumerate(present_labels, start=2):
        m  = report.get(lbl, {})
        f1 = m.get("f1-score", 0)
        fill = FILL_GOOD if f1 >= 0.85 else FILL_WARN if f1 >= 0.70 else FILL_BAD
        _cell(ws2, ri, 1, lbl,                           fill=FILL_LABEL.get(lbl))
        _cell(ws2, ri, 2, round(m.get("precision",0),4), fmt="0.0000")
        _cell(ws2, ri, 3, round(m.get("recall",0),4),    fmt="0.0000")
        _cell(ws2, ri, 4, round(f1,4),                   fill=fill, fmt="0.0000")
        _cell(ws2, ri, 5, int(m.get("support",0)))
        ws2.row_dimensions[ri].height = 16

    tr = len(present_labels) + 2
    _cell(ws2, tr, 1, "Overall (Macro)", bold=True)
    _cell(ws2, tr, 2, round(report.get("macro avg",{}).get("precision",0),4), bold=True, fmt="0.0000")
    _cell(ws2, tr, 3, round(report.get("macro avg",{}).get("recall",0),4),    bold=True, fmt="0.0000")
    _cell(ws2, tr, 4, round(macro_f1,4),                                       bold=True, fmt="0.0000")
    _cell(ws2, tr, 5, int(report.get("macro avg",{}).get("support",0)),        bold=True)

    # F1 bar chart
    chart = BarChart()
    chart.type = "col"; chart.title = "F1-Score per Class"
    chart.y_axis.title = "F1-Score"; chart.x_axis.title = "Posture Class"
    chart.style = 10
    data_ref = Reference(ws2, min_col=4, min_row=1, max_row=len(present_labels)+1)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(present_labels)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "8E44AD"
    chart.width = 18; chart.height = 12
    ws2.add_chart(chart, "G2")

    # ── Sheet 3: Confusion Matrix ───────────────────────────────────────────
    ws3 = wb.create_sheet("Confusion Matrix")
    ws3.cell(row=1, column=1, value="Actual \\ Predicted").font = FONT_BOLD
    ws3.cell(row=1, column=1).alignment = CENTER
    ws3.column_dimensions["A"].width = 18

    for ci, lbl in enumerate(present_labels, start=2):
        _hdr(ws3, 1, ci, lbl, width=15)
        ws3.cell(row=ci, column=1, value=lbl).font      = FONT_BOLD
        ws3.cell(row=ci, column=1).alignment = CENTER
        ws3.cell(row=ci, column=1).border    = THIN

    for ri, row_vals in enumerate(cm, start=2):
        total = row_vals.sum()
        for ci, val in enumerate(row_vals, start=2):
            pct  = val / total if total > 0 else 0
            fill = FILL_GOOD if (ri == ci and pct >= 0.85) \
                   else FILL_WARN if (ri == ci and pct >= 0.70) \
                   else FILL_BAD  if (ri == ci) else None
            c = ws3.cell(row=ri, column=ci, value=int(val))
            c.font = FONT_BOLD if ri == ci else FONT_NORMAL
            c.alignment = CENTER; c.border = THIN
            if fill: c.fill = fill

    # ── Sheet 4: Training History ───────────────────────────────────────────
    ws4 = wb.create_sheet("Training History")
    _hdr(ws4, 1, 1, "Epoch",   width=10)
    _hdr(ws4, 1, 2, "Loss",    width=14)
    _hdr(ws4, 1, 3, "Accuracy",width=14)

    for ri, (ep, loss, acc_h) in enumerate(history, start=2):
        _cell(ws4, ri, 1, ep)
        _cell(ws4, ri, 2, loss, fmt="0.0000")
        _cell(ws4, ri, 3, acc_h, fmt="0.0000")
        ws4.row_dimensions[ri].height = 14

    # loss line chart
    lc = LineChart()
    lc.title = "Training Loss"; lc.y_axis.title = "Loss"; lc.x_axis.title = "Epoch"
    lc.style = 10
    loss_ref = Reference(ws4, min_col=2, min_row=1, max_row=len(history)+1)
    lc.add_data(loss_ref, titles_from_data=True)
    lc.series[0].graphicalProperties.line.solidFill = "E74C3C"
    lc.width = 18; lc.height = 10
    ws4.add_chart(lc, "E2")

    # accuracy line chart
    ac = LineChart()
    ac.title = "Training Accuracy"; ac.y_axis.title = "Accuracy"; ac.x_axis.title = "Epoch"
    ac.style = 10
    acc_ref = Reference(ws4, min_col=3, min_row=1, max_row=len(history)+1)
    ac.add_data(acc_ref, titles_from_data=True)
    ac.series[0].graphicalProperties.line.solidFill = "27AE60"
    ac.width = 18; ac.height = 10
    ws4.add_chart(ac, "E22")

    wb.active = wb["Summary"]
    wb.save(output_path)
    print(f"\n  Results saved: {output_path}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    master            = load_dataset(DATASET_DIR, SHEET_NAME)
    master, feat_cols = prepare_data(master)

    X_train, y_train, X_test, y_test, le = split_and_sequence(master, feat_cols)

    input_size  = X_train.shape[2]
    num_classes = len(le.classes_)

    model, device, history = train_lstm(X_train, y_train, input_size, num_classes)

    # save model
    model_path = os.path.join(OUTPUT_DIR, "lstm_model.pt")
    torch.save({
        "model_state": model.state_dict(),
        "le_classes":  le.classes_.tolist(),
        "feat_cols":   feat_cols,
        "input_size":  input_size,
        "hidden_size": HIDDEN_SIZE,
        "num_layers":  NUM_LAYERS,
        "dropout":     DROPOUT,
        "seq_len":     SEQUENCE_LEN,
        "num_classes": num_classes,
    }, model_path)
    print(f"  Model saved: {model_path}")

    report, cm, present_labels, acc, macro_f1 = evaluate(model, le, X_test, y_test, device)

    out_xlsx = os.path.join(OUTPUT_DIR, "lstm_results.xlsx")
    write_results_excel(out_xlsx, report, cm, present_labels,
                        acc, macro_f1, history, feat_cols)

    print("\nLSTM training complete.")


if __name__ == "__main__":
    main()
