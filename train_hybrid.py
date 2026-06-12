"""
train_hybrid.py  --  Train XGB+LSTM hybrid classifier on sleep posture keypoint data.

Architecture:
    Raw keypoints --> LSTM --> temporal feature vector
                                        |
                              concatenate with raw keypoints
                                        |
                                   XGBoost classifier

Standalone script. Change the CONFIG section at the top, then run:
    python train_hybrid.py

Outputs (saved to OUTPUT_DIR):
    hybrid_model.pkl           -- trained hybrid model, reusable
    hybrid_results.xlsx        -- per-class F1, confusion matrix, comparison chart
"""

# =============================================================================
# CONFIG  --  change these paths each time you use a new dataset
# =============================================================================
DATASET_DIR   = r"C:\SleepPosture\dataset"                  # folder with *_labels.xlsx files
OUTPUT_DIR    = r"C:\SleepPosture\results\XGB+LSTM Training" # where model + results are saved
SHEET_NAME    = "All Frames"                                 # sheet inside each Excel to read
TEST_PERSONS  = ["P05"]                                      # person IDs held out for testing
TRAIN_PERSONS = ["P01", "P02", "P03", "P04"]                # person IDs used for training
                                                             # set both to [] to auto split 80/20

# LSTM feature extractor hyperparameters
SEQUENCE_LEN  = 30       # frames per sequence (30 frames = 3 sec at 10fps)
HIDDEN_SIZE   = 128      # LSTM hidden units
NUM_LAYERS    = 2        # stacked LSTM layers
DROPOUT       = 0.3      # dropout between layers
LSTM_EPOCHS   = 40       # epochs to pretrain LSTM feature extractor
BATCH_SIZE    = 64       # samples per batch
LSTM_LR       = 0.001    # Adam learning rate for LSTM

# XGBoost classifier hyperparameters
N_ESTIMATORS  = 300
MAX_DEPTH     = 6
LEARNING_RATE = 0.1
# =============================================================================

import os
import sys
import glob
import pickle
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

try:
    import torch
    import torch.nn as nn
    from torch.utils.data import DataLoader, TensorDataset
except ImportError:
    print("[ERROR] PyTorch not installed. Run: pip install torch")
    sys.exit(1)

try:
    import xgboost as xgb
except ImportError:
    print("[ERROR] xgboost not installed. Run: pip install xgboost")
    sys.exit(1)

from sklearn.metrics import (classification_report, confusion_matrix,
                             accuracy_score, f1_score)
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split


# =============================================================================
# CONSTANTS
# =============================================================================
LABELS = ["Supine", "Lateral_Left", "Lateral_Right", "Prone", "No_Person"]

FEATURE_SUFFIXES = ["_x", "_y", "_conf"]
DERIVED_FEATURES = ["body_angle", "shoulder_asym", "ear_asym", "movement"]

FILL_HEADER = PatternFill("solid", start_color="2C3E50", end_color="2C3E50")
FILL_GOOD   = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")
FILL_WARN   = PatternFill("solid", start_color="FDEBD0", end_color="FDEBD0")
FILL_BAD    = PatternFill("solid", start_color="FADBD8", end_color="FADBD8")
FILL_LABEL  = {
    "Supine"        : PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3"),
    "Lateral_Left"  : PatternFill("solid", start_color="D6EAF8", end_color="D6EAF8"),
    "Lateral_Right" : PatternFill("solid", start_color="D2E9F7", end_color="D2E9F7"),
    "Prone"         : PatternFill("solid", start_color="FDEBD0", end_color="FDEBD0"),
    "No_Person"     : PatternFill("solid", start_color="EAECEE", end_color="EAECEE"),
}
FONT_HEADER = Font(bold=True, color="FFFFFF", name="Arial", size=9)
FONT_BOLD   = Font(bold=True, name="Arial", size=9)
FONT_NORMAL = Font(name="Arial", size=9)
FONT_TITLE  = Font(bold=True, name="Arial", size=12, color="2C3E50")
THIN        = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)
CENTER      = Alignment(horizontal="center", vertical="center")


# =============================================================================
# LSTM FEATURE EXTRACTOR
# =============================================================================
class LSTMFeatureExtractor(nn.Module):
    """LSTM trained to classify postures — hidden state used as feature vector."""
    def __init__(self, input_size, hidden_size, num_layers, num_classes, dropout):
        super().__init__()
        self.lstm = nn.LSTM(
            input_size, hidden_size, num_layers,
            batch_first=True,
            dropout=dropout if num_layers > 1 else 0.0
        )
        self.dropout = nn.Dropout(dropout)
        self.fc      = nn.Linear(hidden_size, num_classes)

    def forward(self, x):
        out, _ = self.lstm(x)
        feat   = self.dropout(out[:, -1, :])   # last time step = temporal feature
        logits = self.fc(feat)
        return logits, feat                     # return both for hybrid use


# =============================================================================
# DATA LOADING
# =============================================================================
def load_dataset(dataset_dir, sheet_name):
    files = glob.glob(os.path.join(dataset_dir, "*_labels.xlsx"))
    if not files:
        print(f"[ERROR] No *_labels.xlsx files found in: {dataset_dir}")
        sys.exit(1)

    print(f"\nFound {len(files)} Excel file(s):")
    dfs = []
    for f in files:
        print(f"  Loading: {os.path.basename(f)}")
        df = pd.read_excel(f, sheet_name=sheet_name)
        dfs.append(df)

    master = pd.concat(dfs, ignore_index=True)
    print(f"  Total rows loaded: {len(master)}")
    return master


def prepare_data(master):
    if "verified_label" in master.columns:
        master["label"] = master["verified_label"].astype(str).str.strip()
        master["label"] = master["label"].replace({"": None, "nan": None, "None": None})
        master["label"] = master["label"].combine_first(master["auto_label"])
    else:
        master["label"] = master["auto_label"]

    master = master[master["label"].isin(LABELS)].copy()

    if "flag" in master.columns and "verified_label" in master.columns:
        unverified = (master["flag"].astype(str).str.strip() != "") & \
                     (master["verified_label"].astype(str).str.strip().isin(["", "nan", "None"]))
        dropped = unverified.sum()
        master  = master[~unverified].copy()
        if dropped:
            print(f"  Dropped {dropped} unverified flagged rows")

    feat_cols = [c for c in master.columns
                 if any(c.endswith(s) for s in FEATURE_SUFFIXES)
                 or c in DERIVED_FEATURES]
    feat_cols = [c for c in feat_cols if c in master.columns]
    master[feat_cols] = master[feat_cols].fillna(0)

    print(f"  Clean samples: {len(master)} | Features: {len(feat_cols)}")
    print(f"  Label distribution:\n{master['label'].value_counts().to_string()}")
    return master, feat_cols


def make_sequences(df, feat_cols, le, seq_len):
    arr    = df[feat_cols].values.astype(np.float32)
    labels = le.transform(df["label"].values)
    X_list, y_list = [], []
    for i in range(len(arr) - seq_len + 1):
        X_list.append(arr[i : i + seq_len])
        y_list.append(labels[i + seq_len - 1])
    return np.array(X_list, dtype=np.float32), np.array(y_list, dtype=np.int64)


def split_data(master, feat_cols):
    le = LabelEncoder()
    le.fit(LABELS)

    if TRAIN_PERSONS and TEST_PERSONS and "person_id" in master.columns:
        train_df = master[master["person_id"].isin(TRAIN_PERSONS)]
        test_df  = master[master["person_id"].isin(TEST_PERSONS)]
        print(f"\n  Train persons: {TRAIN_PERSONS} ({len(train_df)} rows)")
        print(f"  Test persons:  {TEST_PERSONS} ({len(test_df)} rows)")
    else:
        train_df, test_df = train_test_split(
            master, test_size=0.2, random_state=42, stratify=master["label"]
        )

    X_seq_train, y_seq_train = make_sequences(train_df, feat_cols, le, SEQUENCE_LEN)
    X_seq_test,  y_seq_test  = make_sequences(test_df,  feat_cols, le, SEQUENCE_LEN)

    # raw features aligned to sequence end frames (for XGBoost input)
    raw_train = train_df[feat_cols].values.astype(np.float32)
    raw_test  = test_df[feat_cols].values.astype(np.float32)
    X_raw_train = raw_train[SEQUENCE_LEN - 1:]
    X_raw_test  = raw_test[SEQUENCE_LEN - 1:]

    print(f"  Train sequences: {len(X_seq_train)} | Test sequences: {len(X_seq_test)}")
    return (X_seq_train, y_seq_train,
            X_seq_test,  y_seq_test,
            X_raw_train, X_raw_test, le)


# =============================================================================
# STEP 1 — PRETRAIN LSTM FEATURE EXTRACTOR
# =============================================================================
def pretrain_lstm(X_seq_train, y_seq_train, input_size, num_classes, device):
    print("\nStep 1: Pretraining LSTM feature extractor ...")
    model = LSTMFeatureExtractor(
        input_size, HIDDEN_SIZE, NUM_LAYERS, num_classes, DROPOUT
    ).to(device)

    class_counts = np.bincount(y_seq_train, minlength=num_classes).astype(np.float32)
    class_counts = np.where(class_counts == 0, 1, class_counts)
    weights      = torch.tensor(1.0 / class_counts).to(device)

    criterion = nn.CrossEntropyLoss(weight=weights)
    optimizer = torch.optim.Adam(model.parameters(), lr=LSTM_LR)
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=12, gamma=0.5)

    X_t  = torch.tensor(X_seq_train)
    y_t  = torch.tensor(y_seq_train)
    loader = DataLoader(TensorDataset(X_t, y_t), batch_size=BATCH_SIZE, shuffle=True)

    for epoch in range(1, LSTM_EPOCHS + 1):
        model.train()
        total_loss = 0.0; correct = 0; total = 0
        for xb, yb in loader:
            xb, yb = xb.to(device), yb.to(device)
            optimizer.zero_grad()
            logits, _ = model(xb)
            loss = criterion(logits, yb)
            loss.backward()
            optimizer.step()
            total_loss += loss.item() * len(xb)
            correct    += (logits.argmax(1) == yb).sum().item()
            total      += len(xb)
        scheduler.step()
        if epoch % 10 == 0 or epoch == 1:
            print(f"  Epoch {epoch:3d}/{LSTM_EPOCHS}  "
                  f"loss={total_loss/total:.4f}  acc={correct/total:.4f}")

    print("  LSTM pretraining done.")
    return model


# =============================================================================
# STEP 2 — EXTRACT LSTM FEATURES
# =============================================================================
def extract_lstm_features(lstm_model, X_seq, device, batch_size=256):
    lstm_model.eval()
    features = []
    with torch.no_grad():
        for i in range(0, len(X_seq), batch_size):
            xb = torch.tensor(X_seq[i : i + batch_size]).to(device)
            _, feat = lstm_model(xb)
            features.append(feat.cpu().numpy())
    return np.concatenate(features, axis=0)


# =============================================================================
# STEP 3 — TRAIN XGBOOST ON COMBINED FEATURES
# =============================================================================
def train_xgboost(X_combined_train, y_train, le):
    print("\nStep 3: Training XGBoost on LSTM features + raw keypoints ...")
    y_enc = le.transform(le.inverse_transform(y_train))   # already encoded

    model = xgb.XGBClassifier(
        n_estimators      = N_ESTIMATORS,
        max_depth         = MAX_DEPTH,
        learning_rate     = LEARNING_RATE,
        use_label_encoder = False,
        eval_metric       = "mlogloss",
        random_state      = 42,
    )
    model.fit(X_combined_train, y_train)
    print("  XGBoost training done.")
    return model


# =============================================================================
# EVALUATION
# =============================================================================
def evaluate(xgb_model, le, X_combined_test, y_test_enc):
    y_pred_enc    = xgb_model.predict(X_combined_test)
    y_test_labels = le.inverse_transform(y_test_enc)
    y_pred_labels = le.inverse_transform(y_pred_enc)

    present_labels = sorted(set(y_test_labels) | set(y_pred_labels),
                            key=lambda x: LABELS.index(x) if x in LABELS else 99)

    report   = classification_report(y_test_labels, y_pred_labels,
                                     labels=present_labels, output_dict=True)
    cm       = confusion_matrix(y_test_labels, y_pred_labels, labels=present_labels)
    acc      = accuracy_score(y_test_labels, y_pred_labels)
    macro_f1 = f1_score(y_test_labels, y_pred_labels,
                        average="macro", labels=present_labels)

    print(f"\n  Accuracy:  {acc:.4f}")
    print(f"  Macro F1:  {macro_f1:.4f}")
    print("\n" + classification_report(y_test_labels, y_pred_labels,
                                       labels=present_labels))
    return report, cm, present_labels, acc, macro_f1


# =============================================================================
# EXCEL RESULTS
# =============================================================================
def _hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_HEADER; c.fill = FILL_HEADER
    c.alignment = CENTER; c.border = THIN
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row, col, value, bold=False, fill=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = FONT_BOLD if bold else FONT_NORMAL
    c.alignment = CENTER; c.border = THIN
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    return c


def write_results_excel(output_path, report, cm, present_labels,
                        acc, macro_f1, feat_cols):
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28

    ws["A1"] = "XGB+LSTM Hybrid — Sleep Posture Classification"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = CENTER

    rows = [
        ("Model",              "XGB + LSTM Hybrid"),
        ("Architecture",       "LSTM features → XGBoost classifier"),
        ("Sequence Length",    SEQUENCE_LEN),
        ("LSTM Hidden Size",   HIDDEN_SIZE),
        ("LSTM Layers",        NUM_LAYERS),
        ("LSTM Epochs",        LSTM_EPOCHS),
        ("XGB Estimators",     N_ESTIMATORS),
        ("XGB Max Depth",      MAX_DEPTH),
        ("Train Persons",      ", ".join(TRAIN_PERSONS) if TRAIN_PERSONS else "Auto 80%"),
        ("Test Persons",       ", ".join(TEST_PERSONS)  if TEST_PERSONS  else "Auto 20%"),
        ("Raw Features",       len(feat_cols)),
        ("Combined Features",  f"{len(feat_cols)} + {HIDDEN_SIZE} (LSTM)"),
        ("Overall Accuracy",   f"{acc:.4f}"),
        ("Macro F1",           f"{macro_f1:.4f}"),
    ]
    for ri, (k, v) in enumerate(rows, start=3):
        ws.cell(row=ri, column=1, value=k).font = FONT_BOLD
        ws.cell(row=ri, column=2, value=v).font = FONT_NORMAL
        ws.row_dimensions[ri].height = 16

    # ── Sheet 2: Per-Class Metrics ──────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Class Metrics")
    headers = ["Class", "Precision", "Recall", "F1-Score", "Support"]
    widths  = [18, 13, 13, 13, 13]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        _hdr(ws2, 1, ci, h, width=w)
    ws2.row_dimensions[1].height = 20

    for ri, lbl in enumerate(present_labels, start=2):
        m  = report.get(lbl, {})
        f1 = m.get("f1-score", 0)
        fill = FILL_GOOD if f1 >= 0.85 else FILL_WARN if f1 >= 0.70 else FILL_BAD
        _cell(ws2, ri, 1, lbl,                           fill=FILL_LABEL.get(lbl))
        _cell(ws2, ri, 2, round(m.get("precision",0),4), fmt="0.0000")
        _cell(ws2, ri, 3, round(m.get("recall",0),4),    fmt="0.0000")
        _cell(ws2, ri, 4, round(f1, 4),                  fill=fill, fmt="0.0000")
        _cell(ws2, ri, 5, int(m.get("support", 0)))
        ws2.row_dimensions[ri].height = 16

    tr = len(present_labels) + 2
    _cell(ws2, tr, 1, "Overall (Macro)", bold=True)
    _cell(ws2, tr, 2, round(report.get("macro avg",{}).get("precision",0),4), bold=True, fmt="0.0000")
    _cell(ws2, tr, 3, round(report.get("macro avg",{}).get("recall",0),4),    bold=True, fmt="0.0000")
    _cell(ws2, tr, 4, round(macro_f1, 4),                                      bold=True, fmt="0.0000")
    _cell(ws2, tr, 5, int(report.get("macro avg",{}).get("support",0)),        bold=True)

    # F1 bar chart
    chart = BarChart()
    chart.type = "col"; chart.title = "F1-Score per Class (XGB+LSTM Hybrid)"
    chart.y_axis.title = "F1-Score"; chart.x_axis.title = "Posture Class"
    chart.style = 10
    data_ref = Reference(ws2, min_col=4, min_row=1, max_row=len(present_labels)+1)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(present_labels)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "E67E22"
    chart.width = 18; chart.height = 12
    ws2.add_chart(chart, "G2")

    # ── Sheet 3: Confusion Matrix ───────────────────────────────────────────
    ws3 = wb.create_sheet("Confusion Matrix")
    ws3.cell(row=1, column=1, value="Actual \\ Predicted").font = FONT_BOLD
    ws3.cell(row=1, column=1).alignment = CENTER
    ws3.column_dimensions["A"].width = 18

    for ci, lbl in enumerate(present_labels, start=2):
        _hdr(ws3, 1, ci, lbl, width=15)
        ws3.cell(row=ci, column=1, value=lbl).font      = FONT_BOLD
        ws3.cell(row=ci, column=1).alignment = CENTER
        ws3.cell(row=ci, column=1).border    = THIN

    for ri, row_vals in enumerate(cm, start=2):
        total = row_vals.sum()
        for ci, val in enumerate(row_vals, start=2):
            pct  = val / total if total > 0 else 0
            fill = FILL_GOOD if (ri == ci and pct >= 0.85) \
                   else FILL_WARN if (ri == ci and pct >= 0.70) \
                   else FILL_BAD  if (ri == ci) else None
            c = ws3.cell(row=ri, column=ci, value=int(val))
            c.font = FONT_BOLD if ri == ci else FONT_NORMAL
            c.alignment = CENTER; c.border = THIN
            if fill: c.fill = fill

    wb.active = wb["Summary"]
    wb.save(output_path)
    print(f"\n  Results saved: {output_path}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Device: {device}")

    master            = load_dataset(DATASET_DIR, SHEET_NAME)
    master, feat_cols = prepare_data(master)

    (X_seq_train, y_seq_train,
     X_seq_test,  y_seq_test,
     X_raw_train, X_raw_test, le) = split_data(master, feat_cols)

    input_size  = X_seq_train.shape[2]
    num_classes = len(le.classes_)

    # Step 1 — pretrain LSTM
    lstm_model = pretrain_lstm(X_seq_train, y_seq_train,
                               input_size, num_classes, device)

    # Step 2 — extract LSTM temporal features
    print("\nStep 2: Extracting LSTM temporal features ...")
    lstm_feats_train = extract_lstm_features(lstm_model, X_seq_train, device)
    lstm_feats_test  = extract_lstm_features(lstm_model, X_seq_test,  device)

    # combine LSTM features + raw keypoints
    X_combined_train = np.concatenate([X_raw_train, lstm_feats_train], axis=1)
    X_combined_test  = np.concatenate([X_raw_test,  lstm_feats_test],  axis=1)
    print(f"  Combined feature size: {X_combined_train.shape[1]} "
          f"({len(feat_cols)} raw + {HIDDEN_SIZE} LSTM)")

    # Step 3 — train XGBoost on combined features
    xgb_model = train_xgboost(X_combined_train, y_seq_train, le)

    # save both models
    model_path = os.path.join(OUTPUT_DIR, "hybrid_model.pkl")
    torch.save(lstm_model.state_dict(),
               os.path.join(OUTPUT_DIR, "hybrid_lstm_weights.pt"))
    with open(model_path, "wb") as f:
        pickle.dump({
            "xgb_model"   : xgb_model,
            "le"          : le,
            "feat_cols"   : feat_cols,
            "input_size"  : input_size,
            "hidden_size" : HIDDEN_SIZE,
            "num_layers"  : NUM_LAYERS,
            "dropout"     : DROPOUT,
            "seq_len"     : SEQUENCE_LEN,
            "num_classes" : num_classes,
        }, f)
    print(f"  Model saved: {model_path}")

    report, cm, present_labels, acc, macro_f1 = evaluate(
        xgb_model, le, X_combined_test, y_seq_test
    )

    out_xlsx = os.path.join(OUTPUT_DIR, "hybrid_results.xlsx")
    write_results_excel(out_xlsx, report, cm, present_labels,
                        acc, macro_f1, feat_cols)

    print("\nXGB+LSTM Hybrid training complete.")


if __name__ == "__main__":
    main()
