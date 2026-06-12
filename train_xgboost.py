"""
train_xgboost.py  --  Train XGBoost classifier on sleep posture keypoint data.

Standalone script. Change the CONFIG section at the top, then run:
    python train_xgboost.py

Outputs (saved to OUTPUT_DIR):
    xgboost_model.pkl          -- trained model, reusable
    xgboost_results.xlsx       -- per-class F1, confusion matrix, feature importance chart
"""

# =============================================================================
# CONFIG  --  change these paths each time you use a new dataset
# =============================================================================
DATASET_DIR   = r"C:\SleepPosture\dataset"       # folder containing all (*_labels.xlsx) files
OUTPUT_DIR    = r"C:\SleepPosture\results\XGBoost Training"   # where model + results are saved
SHEET_NAME    = "All Frames"                      # sheet inside each Excel to read
TEST_PERSONS  = ["P05"]                           # person IDs held out for testing
TRAIN_PERSONS = ["P01", "P02", "P03", "P04"]     # person IDs used for training
                                                  # set both to [] to auto split 80/20

# XGBoost hyperparameters
N_ESTIMATORS  = 300
MAX_DEPTH     = 6
LEARNING_RATE = 0.1
# =============================================================================

import os
import sys
import glob
import pickle
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

try:
    import xgboost as xgb
except ImportError:
    print("[ERROR] xgboost not installed. Run: pip install xgboost")
    sys.exit(1)

from sklearn.metrics import (classification_report, confusion_matrix,
                             accuracy_score, f1_score)
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split


# =============================================================================
# CONSTANTS
# =============================================================================
LABELS = ["Supine", "Lateral_Left", "Lateral_Right", "Prone", "No_Person"]

FEATURE_SUFFIXES = ["_x", "_y", "_conf"]
DERIVED_FEATURES = ["body_angle", "shoulder_asym", "ear_asym", "movement"]

FILL_HEADER  = PatternFill("solid", start_color="2C3E50", end_color="2C3E50")
FILL_EVEN    = PatternFill("solid", start_color="F2F3F4", end_color="F2F3F4")
FILL_GOOD    = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")
FILL_WARN    = PatternFill("solid", start_color="FDEBD0", end_color="FDEBD0")
FILL_BAD     = PatternFill("solid", start_color="FADBD8", end_color="FADBD8")
FILL_LABEL   = {
    "Supine"        : PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3"),
    "Lateral_Left"  : PatternFill("solid", start_color="D6EAF8", end_color="D6EAF8"),
    "Lateral_Right" : PatternFill("solid", start_color="D2E9F7", end_color="D2E9F7"),
    "Prone"         : PatternFill("solid", start_color="FDEBD0", end_color="FDEBD0"),
    "No_Person"     : PatternFill("solid", start_color="EAECEE", end_color="EAECEE"),
}
FONT_HEADER  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
FONT_BOLD    = Font(bold=True, name="Arial", size=9)
FONT_NORMAL  = Font(name="Arial", size=9)
FONT_TITLE   = Font(bold=True, name="Arial", size=12, color="2C3E50")
THIN         = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)
CENTER       = Alignment(horizontal="center", vertical="center")


# =============================================================================
# DATA LOADING
# =============================================================================
def load_dataset(dataset_dir, sheet_name):
    files = glob.glob(os.path.join(dataset_dir, "*_labels.xlsx"))
    if not files:
        print(f"[ERROR] No *_labels.xlsx files found in: {dataset_dir}")
        sys.exit(1)

    print(f"\nFound {len(files)} Excel file(s):")
    dfs = []
    for f in files:
        print(f"  Loading: {os.path.basename(f)}")
        df = pd.read_excel(f, sheet_name=sheet_name)
        dfs.append(df)

    master = pd.concat(dfs, ignore_index=True)
    print(f"  Total rows loaded: {len(master)}")
    return master


def prepare_data(master):
    # use verified_label where filled, fall back to auto_label
    if "verified_label" in master.columns:
        master["label"] = master["verified_label"].astype(str).str.strip()
        master["label"] = master["label"].replace({"": None, "nan": None, "None": None})
        master["label"] = master["label"].combine_first(master["auto_label"])
    else:
        master["label"] = master["auto_label"]

    # drop rows with no label or Unknown
    master = master[master["label"].isin(LABELS)].copy()

    # drop unverified flagged rows
    if "flag" in master.columns and "verified_label" in master.columns:
        unverified = (master["flag"].astype(str).str.strip() != "") & \
                     (master["verified_label"].astype(str).str.strip().isin(["", "nan", "None"]))
        dropped = unverified.sum()
        master  = master[~unverified].copy()
        if dropped:
            print(f"  Dropped {dropped} unverified flagged rows")

    # build feature columns
    feat_cols = [c for c in master.columns
                 if any(c.endswith(s) for s in FEATURE_SUFFIXES)
                 or c in DERIVED_FEATURES]
    feat_cols = [c for c in feat_cols if c in master.columns]

    # fill NaN with 0 for missing keypoints
    master[feat_cols] = master[feat_cols].fillna(0)

    print(f"  Clean samples: {len(master)} | Features: {len(feat_cols)}")
    print(f"  Label distribution:\n{master['label'].value_counts().to_string()}")
    return master, feat_cols


def split_data(master, feat_cols):
    if TRAIN_PERSONS and TEST_PERSONS and "person_id" in master.columns:
        train_df = master[master["person_id"].isin(TRAIN_PERSONS)]
        test_df  = master[master["person_id"].isin(TEST_PERSONS)]
        print(f"\n  Train persons: {TRAIN_PERSONS} ({len(train_df)} rows)")
        print(f"  Test persons:  {TEST_PERSONS} ({len(test_df)} rows)")
    else:
        train_df, test_df = train_test_split(master, test_size=0.2,
                                             random_state=42, stratify=master["label"])
        print(f"\n  Auto split 80/20: train={len(train_df)} test={len(test_df)}")

    X_train = train_df[feat_cols].values
    y_train = train_df["label"].values
    X_test  = test_df[feat_cols].values
    y_test  = test_df["label"].values
    return X_train, y_train, X_test, y_test, feat_cols


# =============================================================================
# TRAINING
# =============================================================================
def train_xgboost(X_train, y_train):
    le = LabelEncoder()
    le.fit(LABELS)
    y_enc = le.transform(y_train)

    model = xgb.XGBClassifier(
        n_estimators      = N_ESTIMATORS,
        max_depth         = MAX_DEPTH,
        learning_rate     = LEARNING_RATE,
        use_label_encoder = False,
        eval_metric       = "mlogloss",
        random_state      = 42,
    )
    print("\nTraining XGBoost ...")
    model.fit(X_train, y_enc)
    print("  Done.")
    return model, le


# =============================================================================
# EVALUATION
# =============================================================================
def evaluate(model, le, X_test, y_test):
    y_enc   = le.transform(y_test)
    y_pred  = model.predict(X_test)
    y_pred_labels = le.inverse_transform(y_pred)

    present_labels = sorted(set(y_test) | set(y_pred_labels),
                            key=lambda x: LABELS.index(x) if x in LABELS else 99)

    report = classification_report(y_test, y_pred_labels,
                                   labels=present_labels, output_dict=True)
    cm     = confusion_matrix(y_test, y_pred_labels, labels=present_labels)
    acc    = accuracy_score(y_test, y_pred_labels)
    macro_f1 = f1_score(y_test, y_pred_labels, average="macro", labels=present_labels)

    print(f"\n  Accuracy:  {acc:.4f}")
    print(f"  Macro F1:  {macro_f1:.4f}")
    print("\n" + classification_report(y_test, y_pred_labels, labels=present_labels))
    return report, cm, present_labels, acc, macro_f1, y_pred_labels


# =============================================================================
# EXCEL RESULTS
# =============================================================================
def _hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_HEADER; c.fill = FILL_HEADER
    c.alignment = CENTER; c.border = THIN
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row, col, value, bold=False, fill=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = FONT_BOLD if bold else FONT_NORMAL
    c.alignment = CENTER
    c.border    = THIN
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    return c


def write_results_excel(output_path, report, cm, present_labels,
                        acc, macro_f1, feat_cols, model, le):
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    ws["A1"] = "XGBoost — Sleep Posture Classification"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = CENTER

    rows = [
        ("Model",          "XGBoost"),
        ("N Estimators",   N_ESTIMATORS),
        ("Max Depth",      MAX_DEPTH),
        ("Learning Rate",  LEARNING_RATE),
        ("Train Persons",  ", ".join(TRAIN_PERSONS) if TRAIN_PERSONS else "Auto 80%"),
        ("Test Persons",   ", ".join(TEST_PERSONS)  if TEST_PERSONS  else "Auto 20%"),
        ("Features",       len(feat_cols)),
        ("Overall Accuracy", f"{acc:.4f}"),
        ("Macro F1",       f"{macro_f1:.4f}"),
    ]
    for ri, (k, v) in enumerate(rows, start=3):
        ws.cell(row=ri, column=1, value=k).font  = FONT_BOLD
        ws.cell(row=ri, column=2, value=v).font  = FONT_NORMAL
        ws.row_dimensions[ri].height = 16

    # ── Sheet 2: Per-Class Metrics ──────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Class Metrics")
    headers = ["Class", "Precision", "Recall", "F1-Score", "Support"]
    widths  = [18, 13, 13, 13, 13]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        _hdr(ws2, 1, ci, h, width=w)
    ws2.row_dimensions[1].height = 20

    for ri, lbl in enumerate(present_labels, start=2):
        m = report.get(lbl, {})
        f1 = m.get("f1-score", 0)
        fill = FILL_GOOD if f1 >= 0.85 else FILL_WARN if f1 >= 0.70 else FILL_BAD
        _cell(ws2, ri, 1, lbl,                      fill=FILL_LABEL.get(lbl))
        _cell(ws2, ri, 2, round(m.get("precision",0),4), fmt="0.0000")
        _cell(ws2, ri, 3, round(m.get("recall",0),4),    fmt="0.0000")
        _cell(ws2, ri, 4, round(f1, 4),              fill=fill, fmt="0.0000")
        _cell(ws2, ri, 5, int(m.get("support",0)))
        ws2.row_dimensions[ri].height = 16

    # totals row
    tr = len(present_labels) + 2
    _cell(ws2, tr, 1, "Overall (Macro)", bold=True)
    _cell(ws2, tr, 2, round(report.get("macro avg",{}).get("precision",0),4), bold=True, fmt="0.0000")
    _cell(ws2, tr, 3, round(report.get("macro avg",{}).get("recall",0),4),    bold=True, fmt="0.0000")
    _cell(ws2, tr, 4, round(macro_f1,4),                                       bold=True, fmt="0.0000")
    _cell(ws2, tr, 5, int(report.get("macro avg",{}).get("support",0)),        bold=True)

    # F1 bar chart
    chart = BarChart()
    chart.type       = "col"
    chart.title      = "F1-Score per Class"
    chart.y_axis.title = "F1-Score"
    chart.x_axis.title = "Posture Class"
    chart.style      = 10
    chart.shape      = 4
    data_ref  = Reference(ws2, min_col=4, min_row=1, max_row=len(present_labels)+1)
    cats_ref  = Reference(ws2, min_col=1, min_row=2, max_row=len(present_labels)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "2980B9"
    chart.width  = 18
    chart.height = 12
    ws2.add_chart(chart, f"G2")

    # ── Sheet 3: Confusion Matrix ───────────────────────────────────────────
    ws3 = wb.create_sheet("Confusion Matrix")
    ws3.cell(row=1, column=1, value="Actual \\ Predicted").font = FONT_BOLD
    ws3.cell(row=1, column=1).alignment = CENTER
    ws3.column_dimensions["A"].width = 18

    for ci, lbl in enumerate(present_labels, start=2):
        _hdr(ws3, 1, ci, lbl, width=15)
        ws3.cell(row=ci, column=1, value=lbl).font = FONT_BOLD
        ws3.cell(row=ci, column=1).alignment = CENTER
        ws3.cell(row=ci, column=1).border    = THIN

    for ri, row_vals in enumerate(cm, start=2):
        total = row_vals.sum()
        for ci, val in enumerate(row_vals, start=2):
            pct  = val / total if total > 0 else 0
            fill = FILL_GOOD if (ri == ci and pct >= 0.85) \
                   else FILL_WARN if (ri == ci and pct >= 0.70) \
                   else FILL_BAD  if (ri == ci) \
                   else None
            c = ws3.cell(row=ri, column=ci, value=int(val))
            c.font = FONT_BOLD if ri == ci else FONT_NORMAL
            c.alignment = CENTER; c.border = THIN
            if fill: c.fill = fill

    # ── Sheet 4: Feature Importance ─────────────────────────────────────────
    ws4 = wb.create_sheet("Feature Importance")
    importances = model.feature_importances_
    fi_pairs    = sorted(zip(feat_cols, importances), key=lambda x: x[1], reverse=True)
    top_n       = min(30, len(fi_pairs))

    _hdr(ws4, 1, 1, "Feature",    width=28)
    _hdr(ws4, 1, 2, "Importance", width=14)

    for ri, (feat, imp) in enumerate(fi_pairs[:top_n], start=2):
        _cell(ws4, ri, 1, feat)
        _cell(ws4, ri, 2, round(float(imp), 6), fmt="0.000000")
        ws4.row_dimensions[ri].height = 15

    fi_chart = BarChart()
    fi_chart.type        = "bar"
    fi_chart.title       = f"Top {top_n} Feature Importances"
    fi_chart.y_axis.title = "Importance"
    fi_chart.style       = 10
    fi_data = Reference(ws4, min_col=2, min_row=1, max_row=top_n+1)
    fi_cats = Reference(ws4, min_col=1, min_row=2, max_row=top_n+1)
    fi_chart.add_data(fi_data, titles_from_data=True)
    fi_chart.set_categories(fi_cats)
    fi_chart.series[0].graphicalProperties.solidFill = "27AE60"
    fi_chart.width  = 20
    fi_chart.height = 22
    ws4.add_chart(fi_chart, "D2")

    wb.active = wb["Summary"]
    wb.save(output_path)
    print(f"\n  Results saved: {output_path}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    master              = load_dataset(DATASET_DIR, SHEET_NAME)
    master, feat_cols   = prepare_data(master)
    X_train, y_train, X_test, y_test, feat_cols = split_data(master, feat_cols)

    model, le = train_xgboost(X_train, y_train)

    # save model
    model_path = os.path.join(OUTPUT_DIR, "xgboost_model.pkl")
    with open(model_path, "wb") as f:
        pickle.dump({"model": model, "le": le, "feat_cols": feat_cols}, f)
    print(f"  Model saved: {model_path}")

    report, cm, present_labels, acc, macro_f1, _ = evaluate(model, le, X_test, y_test)

    out_xlsx = os.path.join(OUTPUT_DIR, "xgboost_results.xlsx")
    write_results_excel(out_xlsx, report, cm, present_labels,
                        acc, macro_f1, feat_cols, model, le)

    print("\nXGBoost training complete.")


if __name__ == "__main__":
    main()
